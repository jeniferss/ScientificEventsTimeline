from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def read_excel_file(filename: str, sheet_name: str) -> list[dict] | None:
    workbook: Workbook = load_workbook(filename=filename)
    sheet: Worksheet = workbook[sheet_name]

    max_row: int = sheet.max_row
    max_column: int = sheet.max_column

    rows = list()
    for row in range(2, max_row):
        line = dict()
        for column in range(1, max_column):
            header = str(sheet.cell(row=1, column=column).value).strip().title()
            content = str(sheet.cell(row=row, column=column).value).strip()
            if header and content: line.update({header: content})
        if line: rows.append(line)

    return rows
